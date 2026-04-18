import csv
from io import BytesIO
from tempfile import NamedTemporaryFile

from django import forms
from django.contrib import admin, messages
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.text import slugify
from numbers_parser import Document
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .models import Answer, FeedbackSession, Question, Subject, Teacher


class ExcelUploadForm(forms.Form):
    file = forms.FileField(
        help_text=" .xlsx, .csv we .numbers faýllary goldanylýar."
    )


class ImportTeachersSubjectsMixin:
    import_template_name = "admin/survey/import_form.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import/",
                self.admin_site.admin_view(self.import_view),
                name=f"{self.model._meta.app_label}_{self.model._meta.model_name}_import",
            ),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["import_url"] = (
            f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_import"
        )
        return super().changelist_view(request, extra_context=extra_context)

    def import_view(self, request: HttpRequest) -> HttpResponse:
        form = ExcelUploadForm(request.POST or None, request.FILES or None)

        if request.method == "POST" and form.is_valid():
            file_obj = form.cleaned_data["file"]
            try:
                result_message = self._import_records(file_obj)
            except ValueError as exc:
                self.message_user(request, str(exc), level=messages.ERROR)
            else:
                self.message_user(
                    request,
                    result_message,
                    level=messages.SUCCESS,
                )
                return redirect(
                    f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist"
                )

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "form": form,
            "title": "Excel arkaly import",
            "subtitle": "Mugallym we ders maglumatlaryny yukle",
        }
        return render(request, self.import_template_name, context)

    def _import_records(self, file_obj):
        headers, rows = self._read_rows(file_obj)
        teacher_column = self._resolve_column(
            headers,
            {"teacher", "teacher_name", "full_name", "mugallym", "mugallym_ady"},
        )

        if self.model is Teacher:
            created_teachers = self._import_teachers_only(rows, teacher_column)
            return f"Import tamamlandy. Taze mugallym: {created_teachers}."

        subject_column = self._resolve_column(
            headers,
            {"subject", "subject_name", "name", "ders", "ders_ady"},
        )
        created_teachers, created_subjects, linked_subjects = (
            self._import_teachers_and_subjects(
                rows, teacher_column, subject_column
            )
        )
        return (
            f"Import tamamlandy. "
            f"Taze mugallym: {created_teachers}, "
            f"taze ders: {created_subjects}, "
            f"baglanan ders: {linked_subjects}."
        )

    def _import_teachers_only(self, rows, teacher_column):
        created_teachers = 0

        for row in rows:
            teacher_name = self._clean_value(row.get(teacher_column))
            if not teacher_name:
                continue

            _, teacher_created = Teacher.objects.get_or_create(full_name=teacher_name)
            if teacher_created:
                created_teachers += 1

        return created_teachers

    def _import_teachers_and_subjects(self, rows, teacher_column, subject_column):
        created_teachers = 0
        created_subjects = 0
        linked_subjects = 0

        for row in rows:
            teacher_name = self._clean_value(row.get(teacher_column))
            subject_name = self._clean_value(row.get(subject_column))

            if not teacher_name or not subject_name:
                continue

            teacher, teacher_created = Teacher.objects.get_or_create(
                full_name=teacher_name
            )
            if teacher_created:
                created_teachers += 1

            subject, subject_created = Subject.objects.get_or_create(
                name=subject_name,
                teacher=teacher,
            )
            if subject_created:
                created_subjects += 1
            else:
                linked_subjects += 1

        return created_teachers, created_subjects, linked_subjects

    def _read_rows(self, file_obj):
        filename = file_obj.name.lower()

        if filename.endswith(".xlsx"):
            return self._read_excel_rows(file_obj)
        if filename.endswith(".csv"):
            return self._read_csv_rows(file_obj)
        if filename.endswith(".numbers"):
            return self._read_numbers_file(file_obj)

        raise ValueError("Faýlyň görnüşi .xlsx, .csv ýa-da .numbers bolmaly.")

    def _read_excel_rows(self, file_obj):
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
        worksheet = workbook.active
        row_iter = worksheet.iter_rows(values_only=True)

        try:
            raw_headers = next(row_iter)
        except StopIteration as exc:
            raise ValueError(".xlsx faýly boş.") from exc

        headers = self._normalize_headers(raw_headers)
        rows = []

        for row in row_iter:
            row_data = {}
            for index, header in enumerate(headers):
                value = row[index] if index < len(row) else ""
                row_data[header] = value
            rows.append(row_data)

        return headers, rows

    def _read_csv_rows(self, file_obj):
        file_obj.seek(0)
        decoded_lines = file_obj.read().decode("utf-8-sig").splitlines()
        reader = csv.DictReader(decoded_lines)

        if not reader.fieldnames:
            raise ValueError(".csv faýlynda sütün atlary ýok.")

        headers = self._normalize_headers(reader.fieldnames)
        rows = []

        for raw_row in reader:
            row = {}
            for original_header, normalized_header in zip(reader.fieldnames, headers):
                row[normalized_header] = raw_row.get(original_header, "")
            rows.append(row)

        return headers, rows

    def _read_numbers_file(self, file_obj):
        with NamedTemporaryFile(suffix=".numbers") as temp_file:
            if hasattr(file_obj, "chunks"):
                for chunk in file_obj.chunks():
                    temp_file.write(chunk)
            else:
                temp_file.write(file_obj.read())
            temp_file.flush()

            document = Document(temp_file.name)
            if not document.sheets:
                raise ValueError(".numbers faýlynda sahypa tapylmady.")

            table = document.sheets[0].tables[0] if document.sheets[0].tables else None
            if table is None:
                raise ValueError(".numbers faýlynda tablisa tapylmady.")

            rows = list(table.rows(values_only=True))
            if not rows:
                raise ValueError(".numbers faýly boş.")

            headers = self._normalize_headers(rows[0])
            data_rows = rows[1:]
            normalized_rows = []

            for row in data_rows:
                row_data = {}
                for index, header in enumerate(headers):
                    value = row[index] if index < len(row) else ""
                    row_data[header] = value
                normalized_rows.append(row_data)

            return headers, normalized_rows

    def _normalize_headers(self, raw_headers):
        headers = []
        for index, header in enumerate(raw_headers, start=1):
            cleaned = self._clean_value(header)
            headers.append(cleaned or f"column_{index}")

        if not any(headers):
            raise ValueError("Faýlda sütün atlary ýok.")

        return headers

    def _clean_value(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def _resolve_column(self, columns, candidates):
        normalized = {slugify(column).replace("-", "_"): column for column in columns}
        for candidate in candidates:
            if candidate in normalized:
                return normalized[candidate]
        joined = ", ".join(columns)
        raise ValueError(
            f"Gerekli sütün tapylmady. Bar bolan sütunlar: {joined}."
        )


class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 0
    can_delete = False
    fields = ("question", "score")
    readonly_fields = ("question", "score")
    show_change_link = False

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(Teacher)
class TeacherAdmin(ImportTeachersSubjectsMixin, admin.ModelAdmin):
    change_list_template = "admin/survey/import_changelist.html"
    list_display = ("full_name", "subject_count")
    search_fields = ("full_name", "subjects__name")
    ordering = ("full_name",)

    @admin.display(description="Ders sany")
    def subject_count(self, obj):
        return obj.subjects.count()


@admin.register(Subject)
class SubjectAdmin(ImportTeachersSubjectsMixin, admin.ModelAdmin):
    change_list_template = "admin/survey/import_changelist.html"
    list_display = ("name", "teacher")
    list_filter = ("teacher",)
    search_fields = ("name", "teacher__full_name")
    ordering = ("name", "teacher__full_name")

    class Media:
        css = {"all": ("survey/admin.css",)}


@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ("short_text", "lesson_type")
    list_filter = ("lesson_type",)
    search_fields = ("text",)
    ordering = ("lesson_type", "id")

    @admin.display(description="Sorag")
    def short_text(self, obj):
        return obj.text[:90]


@admin.register(FeedbackSession)
class FeedbackSessionAdmin(admin.ModelAdmin):
    inlines = (AnswerInline,)
    actions = ("export_to_excel",)
    list_display = (
        "teacher",
        "subject",
        "lesson_type",
        "average_score",
        "comment_preview",
        "created_at",
    )
    list_filter = ("lesson_type", "teacher", "subject", "created_at")
    search_fields = (
        "teacher__full_name",
        "subject__name",
        "comment",
        "answers__question__text",
    )
    readonly_fields = ("created_at",)
    autocomplete_fields = ("teacher", "subject")
    date_hierarchy = "created_at"

    @admin.display(description="Ortacha baha")
    def average_score(self, obj):
        scores = list(obj.answers.values_list("score", flat=True))
        if not scores:
            return "-"
        return round(sum(scores) / len(scores), 2)

    @admin.display(description="Bellik")
    def comment_preview(self, obj):
        if not obj.comment:
            return "-"
        return obj.comment[:60]

    @admin.action(description="Saýlanan sessiýalary Excel-e eksport et")
    def export_to_excel(self, request, queryset):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Feedback Sessions"

        headers = [
            "Sessiya ID",
            "Mugallym",
            "Ders",
            "Sapak gornusi",
            "Bellik",
            "Doredilen wagty",
            "Sorag",
            "Baha",
        ]
        worksheet.append(headers)

        rows_written = 0
        sessions = queryset.select_related("teacher", "subject").prefetch_related(
            "answers__question"
        )

        for session in sessions:
            answers = list(session.answers.all())
            if not answers:
                worksheet.append(
                    [
                        session.pk,
                        session.teacher.full_name,
                        session.subject.name,
                        session.get_lesson_type_display(),
                        session.comment,
                        session.created_at.replace(tzinfo=None),
                        "",
                        "",
                    ]
                )
                rows_written += 1
                continue

            for answer in answers:
                worksheet.append(
                    [
                        session.pk,
                        session.teacher.full_name,
                        session.subject.name,
                        session.get_lesson_type_display(),
                        session.comment,
                        session.created_at.replace(tzinfo=None),
                        answer.question.text,
                        answer.score,
                    ]
                )
                rows_written += 1

        for index, _ in enumerate(headers, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = 24

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            'attachment; filename="feedback_sessions_export.xlsx"'
        )

        self.message_user(
            request,
            f"{rows_written} setir Excel faýlyna eksport edildi.",
            level=messages.SUCCESS,
        )
        return response


@admin.register(Answer)
class AnswerAdmin(admin.ModelAdmin):
    list_display = ("session", "question", "score")
    list_filter = ("question__lesson_type", "score")
    search_fields = (
        "session__teacher__full_name",
        "session__subject__name",
        "question__text",
    )
    autocomplete_fields = ("session", "question")
